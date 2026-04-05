#!/usr/bin/env python3
"""
Spreadsheet Formula Processor
Evaluates all formulas in an Excel file using LibreOffice
"""

import json
import sys
import subprocess
import os
import platform
from pathlib import Path
from openpyxl import load_workbook


def configure_calc_macro():
    """Configure LibreOffice macro for formula evaluation if not already set up"""
    if platform.system() == 'Darwin':
        basic_dir = os.path.expanduser('~/Library/Application Support/LibreOffice/4/user/basic/Standard')
    else:
        basic_dir = os.path.expanduser('~/.config/libreoffice/4/user/basic/Standard')
    
    module_path = os.path.join(basic_dir, 'Module1.xba')
    
    if os.path.exists(module_path):
        with open(module_path, 'r') as f:
            if 'RecalculateAndSave' in f.read():
                return True
    
    if not os.path.exists(basic_dir):
        subprocess.run(['soffice', '--headless', '--terminate_after_init'], 
                      capture_output=True, timeout=10)
        os.makedirs(basic_dir, exist_ok=True)
    
    macro_definition = '''<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>'''
    
    try:
        with open(module_path, 'w') as f:
            f.write(macro_definition)
        return True
    except Exception:
        return False


def process_formulas(filepath, wait_time=30):
    """
    Evaluate formulas in Excel file and identify any errors
    
    Args:
        filepath: Path to Excel file
        wait_time: Maximum time to wait for evaluation (seconds)
    
    Returns:
        dict with error locations and counts
    """
    if not Path(filepath).exists():
        return {'error': f'File {filepath} does not exist'}
    
    full_path = str(Path(filepath).absolute())
    
    if not configure_calc_macro():
        return {'error': 'Failed to configure LibreOffice macro'}
    
    command = [
        'soffice', '--headless', '--norestore',
        'vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application',
        full_path
    ]
    
    # Handle timeout command differences between Linux and macOS
    if platform.system() != 'Windows':
        timeout_bin = 'timeout' if platform.system() == 'Linux' else None
        if platform.system() == 'Darwin':
            # Check if gtimeout is available on macOS
            try:
                subprocess.run(['gtimeout', '--version'], capture_output=True, timeout=1, check=False)
                timeout_bin = 'gtimeout'
            except (FileNotFoundError, subprocess.TimeoutExpired):
                pass
        
        if timeout_bin:
            command = [timeout_bin, str(wait_time)] + command
    
    proc_result = subprocess.run(command, capture_output=True, text=True)
    
    if proc_result.returncode != 0 and proc_result.returncode != 124:  # 124 is timeout exit code
        err_output = proc_result.stderr or 'Unknown error during formula evaluation'
        if 'Module1' in err_output or 'RecalculateAndSave' not in err_output:
            return {'error': 'LibreOffice macro not configured properly'}
        else:
            return {'error': err_output}
    
    # Check for Excel errors in the processed file - scan ALL cells
    try:
        workbook = load_workbook(filepath, data_only=True)
        
        error_types = ['#VALUE!', '#DIV/0!', '#REF!', '#NAME?', '#NULL!', '#NUM!', '#N/A']
        error_map = {err: [] for err in error_types}
        error_count = 0
        
        for ws_name in workbook.sheetnames:
            worksheet = workbook[ws_name]
            # Check ALL rows and columns - no limits
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in error_types:
                            if err in cell.value:
                                cell_location = f"{ws_name}!{cell.coordinate}"
                                error_map[err].append(cell_location)
                                error_count += 1
                                break
        
        workbook.close()
        
        # Build result summary
        output = {
            'status': 'success' if error_count == 0 else 'errors_detected',
            'error_count': error_count,
            'error_breakdown': {}
        }
        
        # Add non-empty error categories
        for err_type, cells in error_map.items():
            if cells:
                output['error_breakdown'][err_type] = {
                    'count': len(cells),
                    'cells': cells[:20]  # Show up to 20 locations
                }
        
        # Add formula count for context - also check ALL cells
        wb_with_formulas = load_workbook(filepath, data_only=False)
        formula_total = 0
        for ws_name in wb_with_formulas.sheetnames:
            worksheet = wb_with_formulas[ws_name]
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_total += 1
        wb_with_formulas.close()
        
        output['formula_count'] = formula_total
        
        return output
        
    except Exception as e:
        return {'error': str(e)}


def main():
    if len(sys.argv) < 2:
        print("Usage: python formula_processor.py <excel_file> [timeout_seconds]")
        print("\nEvaluates all formulas in an Excel file using LibreOffice")
        print("\nReturns JSON with error details:")
        print("  - status: 'success' or 'errors_detected'")
        print("  - error_count: Total number of Excel errors found")
        print("  - formula_count: Number of formulas in the file")
        print("  - error_breakdown: Breakdown by error type with locations")
        print("    - #VALUE!, #DIV/0!, #REF!, #NAME?, #NULL!, #NUM!, #N/A")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    wait_time = int(sys.argv[2]) if len(sys.argv) > 2 else 30
    
    output = process_formulas(excel_file, wait_time)
    print(json.dumps(output, indent=2))


if __name__ == '__main__':
    main()
